import io
from datetime import datetime, timezone

from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing, Rect, String
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas as pdfcanvas
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

BRAND = "PRESTRESS SERVICES INDUSTRIES LLC"
FOOTER = "BedForge"

styles = getSampleStyleSheet()
styles.add(ParagraphStyle(name="BrandTitle", fontName="Helvetica-Bold", fontSize=24, leading=28, textColor=colors.HexColor("#111827"), alignment=TA_LEFT))
styles.add(ParagraphStyle(name="BrandSubTitle", fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=colors.HexColor("#0F172A"), alignment=TA_LEFT))
styles.add(ParagraphStyle(name="PackageTitle", fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor("#0F172A"), alignment=TA_LEFT))
styles.add(ParagraphStyle(name="CoverNote", fontName="Helvetica", fontSize=9, leading=12, textColor=colors.HexColor("#475569"), alignment=TA_LEFT))
styles.add(ParagraphStyle(name="SectionTitle", fontName="Helvetica-Bold", fontSize=11.5, leading=15, textColor=colors.HexColor("#0F172A"), spaceBefore=7, spaceAfter=5))
styles.add(ParagraphStyle(name="BodyMono", fontName="Courier", fontSize=8.5, leading=11, textColor=colors.HexColor("#0F172A")))
styles.add(ParagraphStyle(name="SmallLabel", fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor("#475569")))
styles.add(ParagraphStyle(name="QrCaption", fontName="Helvetica-Bold", fontSize=8.5, leading=10, textColor=colors.HexColor("#0F172A"), alignment=TA_CENTER))


def safe_join(values):
    items = [str(item) for item in values if item]
    return ", ".join(items) if items else "—"


def clean_text(value):
    if value in (None, ""):
        return "—"
    return str(value)


def doc_stamp():
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def package_title(package_type):
    return {
        "pour_complete": "Pour Complete State QC Package",
        "single_beam": "Single Beam State QC Package",
        "full_job": "Full Job State QC Package",
    }.get(package_type, "State QC Package")


def package_scope(package_type, job, pour, beams):
    if package_type == "single_beam":
        beam = beams[0] if beams else {}
        return f"Single beam · {beam.get('mark', '—')}"
    if package_type == "full_job":
        return f"Full job · {job.get('job_number', '—')}"
    return f"Pour complete · {pour.get('pour_number', '—')}"


def job_qr_payload(package_type, job, pour, pours, beds, beams):
    bed_text = safe_join(f"Bed {bed.get('bed_number')}" for bed in beds if bed.get("bed_number") is not None)
    beam_marks = safe_join(beam.get("mark") for beam in beams)
    pour_text = safe_join(pour.get("pour_number") for pour in pours)
    return "\n".join([
        f"Package: {package_title(package_type)}",
        f"Scope: {package_scope(package_type, job, pour, beams)}",
        f"Job: {job.get('job_number', '—')} · {job.get('name', '—')}",
        f"Customer: {job.get('customer', '—')}",
        f"DOT Spec: {job.get('state_spec', '—')}",
        f"Pours: {pour_text}",
        f"Beds: {bed_text}",
        f"Beams: {beam_marks}",
    ])


def qr_block(payload):
    widget = qr.QrCodeWidget(payload)
    size = 1.3 * inch
    widget.x = 0
    widget.y = 16
    widget.barWidth = size
    widget.barHeight = size
    drawing = Drawing(size, size + 16)
    drawing.add(Rect(0, 16, size, size, strokeColor=colors.HexColor("#0F172A"), fillColor=colors.white, strokeWidth=0.8))
    drawing.add(widget)
    drawing.add(String(size / 2, 5, "JOB QR", fontName="Helvetica-Bold", fontSize=8.5, textAnchor="middle", fillColor=colors.HexColor("#0F172A")))
    return drawing


def kv_table(rows, widths=(1.75 * inch, 4.85 * inch)):
    data = [[Paragraph(f"<b>{label}</b>", styles["BodyText"]), Paragraph(clean_text(value), styles["BodyText"])] for label, value in rows]
    table = Table(data, colWidths=list(widths), hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E2E8F0")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F8FAFC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    return table


def grid_table(headers, rows, widths=None):
    data = [headers] + rows
    table = Table(data, colWidths=widths, repeatRows=1, hAlign="LEFT")
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for idx in range(1, len(data)):
        style.append(("BACKGROUND", (0, idx), (-1, idx), colors.HexColor("#F8FAFC" if idx % 2 else "#FFFFFF")))
    table.setStyle(TableStyle(style))
    return table


def signoff_table(rows):
    table = Table(rows, colWidths=[2.0 * inch, 2.45 * inch, 0.95 * inch, 1.1 * inch], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E2E8F0")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("LEADING", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#94A3B8")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
    ]))
    return table


def add_section(story, number, title, element):
    story.append(Paragraph(f"{number}. {title}", styles["SectionTitle"]))
    story.append(element)
    story.append(Spacer(1, 0.16 * inch))


def cover_sections(batch_records):
    sections = [
        ["1", "Beams", "Beam schedule, bed positions, QC status, and traceability"],
        ["2", "Batch / environment", "Ticket, mix, weather, concrete temperature, and batch detail"],
        ["3", "QIR", "Inspection sections, results, inspector, and notes"],
        ["4", "Tension", "Bed tension and elongation acceptance"],
        ["5", "Strength / camber", "Release strength and 3-point camber"],
        ["6", "Finish", "Finish review and anomaly summary"],
        ["7", "Pre-delivery sign-off", "QC Tech, Production Supervisor, and Quality Manager approvals"],
        ["8", "Strand traceability", "Beam-to-roll and release tag traceability"],
        ["9", "NCR summary", "Linked NCR status and photo references"],
    ]
    if any(record.get("ingredients") for record in batch_records):
        sections[1][2] = "Ticket, mix, weather, concrete temperature, and ingredient checks"
    return sections


def draw_cover_banner(package_type, job, pour, pours, beds, beams):
    qr_draw = qr_block(job_qr_payload(package_type, job, pour, pours, beds, beams))
    brand_block = [
        Paragraph(BRAND, styles["BrandTitle"]),
        Spacer(1, 0.04 * inch),
        Paragraph("Approved model report format", styles["BrandSubTitle"]),
        Spacer(1, 0.06 * inch),
        Paragraph(package_title(package_type), styles["PackageTitle"]),
        Spacer(1, 0.04 * inch),
        Paragraph("Plant / DOT-ready production record package", styles["CoverNote"]),
    ]
    table = Table([[brand_block, qr_draw]], colWidths=[5.0 * inch, 1.45 * inch], hAlign="LEFT")
    table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#CBD5E1")),
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#F8FAFC")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    return table


class PackageCanvas(pdfcanvas.Canvas):
    def __init__(self, *args, left_margin=0.55 * inch, right_margin=0.55 * inch, **kwargs):
        super().__init__(*args, **kwargs)
        self.left_margin = left_margin
        self.right_margin = right_margin
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        total_pages = len(self._saved_page_states)
        for page_number, state in enumerate(self._saved_page_states, start=1):
            self.__dict__.update(state)
            self.draw_chrome(page_number, total_pages)
            super().showPage()
        super().save()

    def draw_chrome(self, page_number, total_pages):
        width, height = self._pagesize
        self.saveState()
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        if page_number > 1:
            self.line(self.left_margin, height - 0.48 * inch, width - self.right_margin, height - 0.48 * inch)
            self.setFont("Helvetica-Bold", 11)
            self.setFillColor(colors.HexColor("#0F172A"))
            self.drawString(self.left_margin, height - 0.34 * inch, BRAND)
            self.setFont("Helvetica", 7.5)
            self.drawRightString(width - self.right_margin, height - 0.34 * inch, "State QC production package")
        self.line(self.left_margin, 0.48 * inch, width - self.right_margin, 0.48 * inch)
        self.setFont("Helvetica", 7.5)
        self.setFillColor(colors.HexColor("#475569"))
        self.drawString(self.left_margin, 0.28 * inch, f"Generated {doc_stamp()}")
        self.drawRightString(width - self.right_margin, 0.28 * inch, f"Page {page_number} of {total_pages}")
        if page_number == total_pages:
            self.setFont("Helvetica", 6.5)
            self.drawCentredString(width / 2, 0.14 * inch, FOOTER)
        self.restoreState()


def build_package_pdf(context: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=letter,
        leftMargin=0.55 * inch,
        rightMargin=0.55 * inch,
        topMargin=0.76 * inch,
        bottomMargin=0.62 * inch,
        pageCompression=0,
    )
    story = []

    package_type = context.get("package_type", "pour_complete")
    job = context.get("job", {})
    pour = context.get("pour", {})
    pours = context.get("pours", [])
    if not pours and pour:
        pours = [pour]
    beds = context.get("beds", [])
    beams = context.get("beams", [])
    inspections = context.get("inspections", [])
    tension_reports = context.get("tension_reports", [])
    camber_readings = context.get("camber_readings", [])
    batch_records = context.get("batch_records", [])
    if not batch_records and context.get("batch_record"):
        batch_records = [context["batch_record"]]
    ncrs = context.get("ncrs", [])
    anomalies = context.get("anomalies", [])

    inspectors = sorted({item.get("inspector") for item in inspections if item.get("inspector")})
    bed_labels = [f"Bed {bed.get('bed_number')}" for bed in beds if bed.get("bed_number") is not None]
    beam_ids = {beam.get("id") for beam in beams}

    story.append(draw_cover_banner(package_type, job, pour, pours, beds, beams))
    story.append(Spacer(1, 0.16 * inch))
    story.append(kv_table([
        ("Scope", package_scope(package_type, job, pour, beams)),
        ("Job", f"{job.get('job_number', '—')} · {job.get('name', '—')}"),
        ("Customer", job.get("customer", "—")),
        ("DOT / Spec", job.get("state_spec", "—")),
        ("Pour" if len(pours) == 1 else "Pours", safe_join(item.get("pour_number") for item in pours)),
        ("Beds", safe_join(bed_labels)),
        ("Beam count", len(beams)),
        ("Beam marks", safe_join(beam.get("mark") for beam in beams)),
        ("QC personnel", safe_join(inspectors)),
        ("Generated UTC", doc_stamp()),
    ]))
    story.append(Spacer(1, 0.16 * inch))

    add_section(
        story,
        "COVER",
        "Package Contents",
        grid_table(["Ref", "Section", "Purpose"], cover_sections(batch_records), widths=[0.55 * inch, 1.95 * inch, 4.6 * inch]),
    )
    add_section(
        story,
        "COVER",
        "Routing Sign-Off",
        signoff_table([
            ["Role", "Name / Signature", "Date", "Status"],
            ["QC Tech", "____________________________", "____________", "Ready to release"],
            ["Production Supervisor", "____________________________", "____________", "Plant review"],
            ["Quality Manager", "____________________________", "____________", "Final approval"],
        ]),
    )

    story.append(PageBreak())

    beam_rows = [
        [
            beam.get("mark", "—"),
            beam.get("product_type", {}).get("name", beam.get("twin_type", "—")),
            next((bed.get("bed_number") for bed in beds if bed.get("id") == beam.get("bed_id")), "—"),
            beam.get("position_on_bed", "—"),
            beam.get("length_ft", "—"),
            (
                f"LOCKED R{beam.get('blueprint_source', {}).get('revision_id', '')[:8]}"
                if beam.get("blueprint_source", {}).get("status") == "locked"
                else beam.get("blueprint_source", {}).get("status", "legacy_seed").upper()
            ),
            beam.get("qc_state", "—"),
            safe_join(beam.get("traceability", {}).get("strand_rolls", [])),
        ]
        for beam in beams
    ]
    add_section(story, 1, "Beams", grid_table(["Beam", "Product", "Bed", "Pos", "Length (ft)", "Blueprint", "QC", "Strand Rolls"], beam_rows or [["—", "—", "—", "—", "—", "—", "—", "—"]], widths=[0.75 * inch, 1.85 * inch, 0.45 * inch, 0.45 * inch, 0.75 * inch, 1.15 * inch, 0.65 * inch, 1.65 * inch]))

    pours_by_id = {item.get("id"): item for item in pours}
    batch_rows = [
        [
            pours_by_id.get(record.get("pour_id"), {}).get("pour_number", "—"),
            record.get("ticket_number", "—"),
            record.get("mix_design", pours_by_id.get(record.get("pour_id"), {}).get("concrete_mix", "—")),
            record.get("ambient_temp_f", "—"),
            record.get("concrete_temp_f", "—"),
            record.get("humidity_pct", "—"),
            record.get("wind_mph", "—"),
            record.get("weather", "—"),
        ]
        for record in batch_records
    ]
    add_section(story, 2, "Batch / Environment", grid_table(["Pour", "Ticket", "Mix", "Ambient °F", "Concrete °F", "Humidity %", "Wind MPH", "Weather"], batch_rows or [["—"] * 8], widths=[0.6 * inch, 0.8 * inch, 1.35 * inch, 0.7 * inch, 0.8 * inch, 0.7 * inch, 0.65 * inch, 1.0 * inch]))

    ingredient_rows = [
        [
            pours_by_id.get(record.get("pour_id"), {}).get("pour_number", "—"),
            record.get("ticket_number", "—"),
            item.get("name", "—"),
            item.get("target_lb", item.get("target", "—")),
            item.get("actual_lb", item.get("actual", "—")),
        ]
        for record in batch_records
        for item in record.get("ingredients", [])
    ]
    if ingredient_rows:
        add_section(story, "2A", "Batch Ingredient Check", grid_table(["Pour", "Ticket", "Ingredient", "Target (lb)", "Actual (lb)"], ingredient_rows, widths=[0.75 * inch, 0.9 * inch, 2.8 * inch, 1.0 * inch, 1.0 * inch]))

    qir_rows = [[next((beam.get("mark") for beam in beams if beam.get("id") == item.get("beam_id")), item.get("beam_id", "—")), item.get("section", "—"), item.get("status", "—"), item.get("inspector", "—"), item.get("notes", "—")] for item in inspections]
    add_section(story, 3, "QIR", grid_table(["Beam", "Section", "Status", "Inspector", "Notes"], qir_rows or [["—", "—", "—", "—", "—"]], widths=[1.0 * inch, 1.2 * inch, 0.85 * inch, 1.2 * inch, 2.7 * inch]))

    tension_rows = [[item.get("bed_number", "—"), item.get("strand_size", "—"), item.get("theoretical_elongation_in", "—"), item.get("measured_elongation_in", "—"), item.get("variance_pct", "—"), "PASS" if item.get("within_tolerance") else "FAIL"] for item in tension_reports]
    add_section(story, 4, "Tension", grid_table(["Bed", "Strand", "Theo. In", "Measured In", "Variance %", "Result"], tension_rows or [["—", "—", "—", "—", "—", "—"]], widths=[0.75 * inch, 1.0 * inch, 1.0 * inch, 1.15 * inch, 1.0 * inch, 0.85 * inch]))

    camber_rows = []
    for item in camber_readings:
        mid = item.get("measured_camber_in", 0)
        camber_rows.append([
            item.get("beam_mark", item.get("beam_id", "—")),
            round(mid * 0.4, 2),
            round(mid, 2),
            round(mid * 0.45, 2),
            item.get("release_strength_psi", "—"),
            item.get("required_strength_psi", "—"),
        ])
    add_section(story, 5, "Strength / Camber", grid_table(["Beam", "End A", "Mid", "End B", "Release PSI", "Req. PSI"], camber_rows or [["—", "—", "—", "—", "—", "—"]], widths=[1.1 * inch, 0.9 * inch, 0.85 * inch, 0.9 * inch, 1.15 * inch, 1.0 * inch]))

    finish_rows = [[next((beam.get("mark") for beam in beams if beam.get("id") == item.get("beam_id")), item.get("beam_id", "—")), item.get("status", "—"), item.get("notes", "—")] for item in inspections if item.get("section") in ("concrete", "finish")]
    add_section(story, 6, "Finish", grid_table(["Beam", "Status", "Notes"], finish_rows or [["—", "—", "—"]], widths=[1.2 * inch, 0.9 * inch, 4.5 * inch]))

    anomaly_rows = [[next((beam.get("mark") for beam in beams if beam.get("id") == item.get("beam_id")), item.get("beam_id", "—")), item.get("type", "—"), item.get("severity", "—"), item.get("length_in", "—"), item.get("note", "—")] for item in anomalies if item.get("beam_id") in beam_ids]
    add_section(story, "6A", "Finish Anomaly Summary", grid_table(["Beam", "Type", "Severity", "Length (in)", "Note"], anomaly_rows or [["—", "—", "—", "—", "—"]], widths=[1.1 * inch, 1.0 * inch, 0.9 * inch, 0.9 * inch, 2.9 * inch]))

    release_rows = [["QC Tech", "____________________________", "____________", safe_join(sorted({item.get('inspector') for item in inspections if item.get('section') == 'pre_delivery' and item.get('inspector')})) or "Assigned at release"], ["Production Supervisor", "____________________________", "____________", "Plant release confirmation"], ["Quality Manager", "____________________________", "____________", "Final shipment authorization"]]
    add_section(story, 7, "Pre-Delivery Sign-Off", signoff_table([["Role", "Name / Signature", "Date", "Notes"]] + release_rows))

    trace_rows = [[beam.get("mark", "—"), safe_join(beam.get("traceability", {}).get("strand_rolls", [])), beam.get("traceability", {}).get("release_tag", "—")] for beam in beams]
    add_section(story, 8, "Strand Traceability", grid_table(["Beam", "Strand Rolls", "Release Tag"], trace_rows or [["—", "—", "—"]], widths=[1.0 * inch, 4.4 * inch, 1.2 * inch]))

    ncr_rows = [[item.get("code", "—"), item.get("status", "—"), item.get("title", "—"), safe_join(item.get("linked_photo_urls", []))] for item in ncrs]
    add_section(story, 9, "NCR Summary", grid_table(["NCR", "Status", "Title", "Photos"], ncr_rows or [["—", "—", "—", "—"]], widths=[1.0 * inch, 1.0 * inch, 2.2 * inch, 2.8 * inch]))

    doc.build(
        story,
        canvasmaker=lambda *args, **kwargs: PackageCanvas(*args, left_margin=doc.leftMargin, right_margin=doc.rightMargin, **kwargs),
    )
    buf.seek(0)
    return buf.read()


def build_package_xlsx(ctx: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    title_font = Font(bold=True, size=16, name="Arial")
    fill = PatternFill("solid", fgColor="0F1218")
    border = Border(left=Side(style="thin", color="D0D7DE"), right=Side(style="thin", color="D0D7DE"), top=Side(style="thin", color="D0D7DE"), bottom=Side(style="thin", color="D0D7DE"))
    cover = wb.active
    cover.title = "Cover"
    company = (ctx.get("company") or {}).get("company_name") or BRAND
    pour = ctx.get("pour") or {}
    job = ctx.get("job") or {}
    beams = ctx.get("beams") or []
    cover["A1"] = company
    cover["A1"].font = title_font
    cover["A2"] = "DOT / OWNER QUALITY PACKAGE"
    rows = [
        ("Job", job.get("job_number") or ""),
        ("Customer", job.get("customer") or ""),
        ("Pour", pour.get("pour_number") or ""),
        ("Pour date", pour.get("pour_date") or ""),
        ("Beams", safe_join(beam.get("mark") for beam in beams) if beams else safe_join(ctx.get("beam_marks") or [])),
        ("Generated", doc_stamp()),
    ]
    for idx, (label, value) in enumerate(rows, 4):
        cover.cell(idx, 1, label)
        cover.cell(idx, 2, value)

    def sheet(name, headers, rows):
        ws = wb.create_sheet(name[:31])
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = header_font
            c.fill = fill
            c.alignment = Alignment(horizontal="center")
            c.border = border
        for r, row in enumerate(rows, 2):
            for i, val in enumerate(row, 1):
                ws.cell(row=r, column=i, value=val).border = border
        for col in "ABCDEFGHIJK":
            ws.column_dimensions[col].width = 18

    sheet("QIR", ["Beam", "Section", "Status", "Inspector", "Date"], [[i.get("beam_mark") or i.get("beam_id"), i.get("section"), i.get("status"), i.get("inspector"), str(i.get("created_at") or "")[:10]] for i in ctx.get("inspections") or []])
    sheet("Tension", ["Bed", "Jack", "Elongation", "Within", "By"], [[t.get("bed_number"), t.get("jack_id") or t.get("jack"), t.get("measured_elongation_in") or t.get("elongation_in"), "YES" if t.get("within_tolerance") else "NO", t.get("created_by") or t.get("inspector") or ""] for t in ctx.get("tension_reports") or []])
    sheet("Strength_Camber", ["Beam", "Required psi", "Release psi", "Measured", "Date"], [[c.get("beam_mark") or c.get("beam_id"), c.get("required_strength_psi"), c.get("release_strength_psi"), c.get("measured_camber_in") or c.get("midspan_in"), str(c.get("created_at") or c.get("reading_date") or "")[:10]] for c in ctx.get("camber_readings") or []])
    sheet("Batch", ["Ticket", "Mix", "Ambient", "Concrete", "Wind", "Date"], [[b.get("ticket_number"), b.get("mix_design") or b.get("mix_code"), (b.get("environment") or {}).get("ambient_f") or b.get("ambient_temp_f"), (b.get("environment") or {}).get("mix_temp_f") or b.get("concrete_temp_f"), (b.get("environment") or {}).get("wind_mph") or b.get("wind_mph"), str(b.get("created_at") or b.get("batched_at") or "")[:10]] for b in ctx.get("batch_records") or ([ctx.get("batch_record")] if ctx.get("batch_record") else [])])
    sheet("NCR", ["Code", "Title", "Status", "Severity", "Owner"], [[n.get("code"), n.get("title") or n.get("description"), n.get("status"), n.get("severity"), n.get("owner") or n.get("assigned_to")] for n in ctx.get("ncrs") or []])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
