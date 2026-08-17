"""Excel export producing close digital versions of the plant QC forms."""
import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="12151C", end_color="12151C", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=12, name="Arial")
TITLE_FONT = Font(bold=True, size=16, name="Arial")
LABEL_FONT = Font(bold=True, size=10, name="Arial")
THIN = Side(style="thin", color="999999")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _title(ws, text, subtitle="", context=None):
    company = (context or {}).get("company_name") or "PRESTRESS SERVICES INDUSTRIES LLC"
    ws["A1"] = company
    ws["A1"].font = TITLE_FONT
    ws["A2"] = text
    ws["A2"].font = Font(bold=True, size=13, name="Arial")
    if subtitle:
        ws["A3"] = subtitle
        ws["A3"].font = Font(italic=True, size=10, name="Arial")
    ws["F1"] = "Generated:"
    ws["G1"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


def _kv_table(ws, start_row, rows):
    r = start_row
    for label, value in rows:
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = LABEL_FONT
        c1.border = BORDER
        c2 = ws.cell(row=r, column=2, value=value)
        c2.border = BORDER
        r += 1
    return r


def _header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER


def build_qir(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "QIR"
    _title(ws, "QUALITY INSPECTION REPORT (QIR 2026.6.1)", context=context)
    beam = context.get("beam", {})
    job = context.get("job", {})
    r = _kv_table(ws, 5, [
        ("Job Number", job.get("job_number", "")),
        ("Job Name", job.get("name", "")),
        ("Customer", job.get("customer", "")),
        ("Beam Mark", beam.get("mark", "")),
        ("Product Type", context.get("product_type_name", "")),
        ("Length (ft)", beam.get("length_ft", "")),
        ("QC State", beam.get("qc_state", "")),
    ])
    r += 1
    _header_row(ws, r, ["Section", "Status", "Inspector", "Notes", "Date"])
    r += 1
    for insp in context.get("inspections", []):
        ws.cell(row=r, column=1, value=insp.get("section", "")).border = BORDER
        ws.cell(row=r, column=2, value=insp.get("status", "")).border = BORDER
        ws.cell(row=r, column=3, value=insp.get("inspector", "")).border = BORDER
        ws.cell(row=r, column=4, value=insp.get("notes", "")).border = BORDER
        ws.cell(row=r, column=5, value=str(insp.get("created_at", ""))[:10]).border = BORDER
        r += 1
    _widths(ws)
    return _save(wb)


def build_tension(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tension Report"
    _title(ws, "STRAND TENSION REPORT", context=context)
    _header_row(ws, 5, ["Bed", "Strand Size", "Area (in²)", "Jack Force (kip)",
                        "Bed Len (ft)", "Theo. Elong (in)", "Meas. Elong (in)",
                        "Variance %", "Result"])
    r = 6
    for t in context.get("tension_reports", []):
        vals = [
            t.get("bed_number", ""), t.get("strand_size", ""),
            t.get("strand_area_in2", ""), t.get("jacking_force_kip", ""),
            t.get("bed_length_ft", ""), t.get("theoretical_elongation_in", ""),
            t.get("measured_elongation_in", ""), t.get("variance_pct", ""),
            "PASS" if t.get("within_tolerance") else "FAIL",
        ]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=1 + i, value=v).border = BORDER
        r += 1
    _widths(ws)
    return _save(wb)


def build_camber(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Camber & Strength"
    _title(ws, "CAMBER / RELEASE STRENGTH SHEET", context=context)
    _header_row(ws, 5, ["Beam Mark", "Design (in)", "Marked End (in)", "Midspan (in)",
                        "Unmarked End (in)", "Required Str (psi)", "Release Str (psi)", "Date"])
    r = 6
    for c in context.get("camber_readings", []):
        vals = [
            c.get("beam_mark", ""),
            c.get("design_camber_in", ""),
            c.get("marked_end_in", ""),
            c.get("midspan_in", c.get("measured_camber_in", "")),
            c.get("unmarked_end_in", ""),
            c.get("required_strength_psi", ""),
            c.get("release_strength_psi", ""),
            str(c.get("reading_date", ""))[:10],
        ]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=1 + i, value=v).border = BORDER
        r += 1
    _widths(ws)
    return _save(wb)


def build_crackmap(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Crack Map"
    _title(ws, "CRACK MAP / ANOMALY LOG", context=context)
    _header_row(ws, 5, ["Beam Mark", "Type", "Severity", "Pos X", "Pos Y", "Pos Z",
                        "Length (in)", "Note", "Date"])
    r = 6
    for a in context.get("anomalies", []):
        pos = a.get("position", {})
        vals = [a.get("beam_mark", ""), a.get("type", ""), a.get("severity", ""),
                pos.get("x", ""), pos.get("y", ""), pos.get("z", ""),
                a.get("length_in", ""), a.get("note", ""), str(a.get("created_at", ""))[:10]]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=1 + i, value=v).border = BORDER
        r += 1
    _widths(ws)
    return _save(wb)


def build_finish(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Finish Sheet"
    _title(ws, "FINISH SHEET — POST-POUR", context=context)
    _header_row(ws, 5, ["Beam Mark", "Marked End ID", "Strand Flush", "Recessed", "Grouted",
                        "Hardware", "Surface", "Surface Pass", "Status", "Inspector", "Date"])
    r = 6
    for s in context.get("finish_sheets", []):
        vals = [
            s.get("beam_mark", ""),
            s.get("marked_end_id", ""),
            "YES" if s.get("strand_cut_flush") else "NO",
            "YES" if s.get("strand_recessed") else "NO",
            "YES" if s.get("strand_grouted") else "NO",
            "YES" if s.get("hardware_complete") else "NO",
            s.get("surface_finish", ""),
            "YES" if s.get("surface_pass") else "NO",
            s.get("status", ""),
            s.get("inspector", ""),
            str(s.get("created_at", ""))[:10],
        ]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=1 + i, value=v).border = BORDER
        r += 1
    _widths(ws)
    return _save(wb)


def build_pre_delivery(context: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pre-Delivery"
    _title(ws, "PRE-DELIVERY / RELEASE", context=context)
    _header_row(ws, 5, ["Beam Mark", "Truck", "Destination", "Load Pos",
                        "QC", "Production", "Carrier", "Released", "Date"])
    r = 6
    for rec in context.get("pre_delivery", []):
        vals = [
            rec.get("beam_mark", ""),
            rec.get("truck_number", ""),
            rec.get("destination", ""),
            rec.get("load_position", ""),
            rec.get("qc_signoff", ""),
            rec.get("production_signoff", ""),
            rec.get("carrier_signoff", ""),
            "YES" if rec.get("released") else "NO",
            str(rec.get("created_at", ""))[:10],
        ]
        for i, v in enumerate(vals):
            ws.cell(row=r, column=1 + i, value=v).border = BORDER
        r += 1
    _widths(ws)
    return _save(wb)


def _widths(ws):
    for col in "ABCDEFGHIJK":
        ws.column_dimensions[col].width = 20


def _save(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


BUILDERS = {
    "qir": ("build_qir", "QIR_2026.6.1"),
    "tension": ("build_tension", "Tension_Report"),
    "camber": ("build_camber", "Camber_Strength"),
    "crackmap": ("build_crackmap", "Crack_Map"),
    "finish": ("build_finish", "Finish_Sheet"),
    "pre_delivery": ("build_pre_delivery", "Pre_Delivery"),
}
